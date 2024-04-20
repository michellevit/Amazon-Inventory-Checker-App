import sys
import openpyxl
import logging

# Setting up logging to print to standard output as well
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[logging.StreamHandler(sys.stdout)])

def calculate_inventory(file_path):
    logging.info(f"Attempting to open the workbook: {file_path}")
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)  # Ensure formulas are evaluated
        sheet = workbook['Line Items'] if 'Line Items' in workbook.sheetnames else None
        
        if sheet is None:
            logging.error("The 'Line Items' sheet does not exist in the workbook.")
            sys.exit(2)  # Exit with error code 2 if the sheet is missing

        logging.info("Processing rows in the 'Line Items' workbook...")
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            logging.info(row)  # Log each row's data to understand what's being processed

        workbook.save(file_path)
        logging.info("Workbook processed and saved successfully.")
    except Exception as e:
        logging.error(f"Failed to process the workbook: {str(e)}")
        sys.exit(1)  # Exit with error code 1 for general errors

if __name__ == "__main__":
    if len(sys.argv) != 2:
        logging.error("Usage: python calculate_inventory.py <path_to_excel_file>")
        sys.exit(1)
    
