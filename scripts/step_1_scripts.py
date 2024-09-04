# step_1_scripts.py

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import os
import pyperclip
from tkinter import messagebox
import xlrd


def convert_xls_to_xlsx(original_path, currency, processing_dir):
    # Read the original workbook using xlrd
    book_xls = xlrd.open_workbook(original_path, formatting_info=True)
    sheet_xls = book_xls.sheet_by_name('Line Items')
    # Create a new workbook using openpyxl
    book_xlsx = Workbook()
    sheet_xlsx = book_xlsx.active
    sheet_xlsx.title = 'Line Items'
    # Copy data from the old xls file to the new xlsx file
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            cell_value = sheet_xls.cell_value(row, col)
            cell_type = sheet_xls.cell_type(row, col)
            if cell_type == xlrd.XL_CELL_DATE:
                date_value = xlrd.xldate.xldate_as_datetime(cell_value, book_xls.datemode)
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = date_value
            else:
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = cell_value
    # filename_without_extension, _ = os.path.splitext(filename)
    capitalized_currency = currency.upper()
    new_filename = f"Amazon-Orders-{capitalized_currency}-Processing.xlsx"
    new_file_path = os.path.join(processing_dir, new_filename)
    # Save the new workbook in the same directory as the original
    book_xlsx.save(new_file_path)
    # Return the new workbook object
    return book_xlsx, new_filename, new_file_path


def create_new_file(original_path, currency, processing_dir):
    workbook = load_workbook(original_path)
    if 'Line Items' not in workbook.sheetnames:
        raise ValueError("No 'Line Items' sheet found in the workbook.")
    original_sheet = workbook['Line Items']
    new_workbook = load_workbook()
    new_sheet = new_workbook.active
    new_sheet.title = 'Line Items'
    for row in original_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
    capitalized_currency = currency.upper()
    new_filename = f"Amazon-Orders-{capitalized_currency}-Processing.xlsx"
    new_file_path = os.path.join(processing_dir, new_filename)
    new_workbook.save(new_file_path)
    return new_workbook, new_filename, new_file_path


def check_file_valid(workbook, currency):
    capitalized_currency = currency.upper()
    if 'Line Items' not in workbook.sheetnames:
        raise ValueError("The file entered is not valid (it does not have a 'Line Items' tab).")
    sheet = workbook['Line Items']
    if sheet['A4'].value:
        if capitalized_currency not in sheet['AF4'].value:
            raise ValueError(f"Please enter the {capitalized_currency} file into the correct input.")
        return True
    else:
        return False


def cancel_orders_below_min(workbook, min_order_value, new_file_path, po_value_dict, orders_to_cancel_array):
    sheet = workbook['Line Items']
    current_currency_orders = []
    for row in range(4, sheet.max_row + 1):
        if not sheet[f'A{row}'].value:
            break
        order_number = sheet[f'A{row}'].value
        if order_number not in current_currency_orders:
            current_currency_orders.append(order_number)
        quantity_ordered = sheet[f'J{row}'].value # Type = float
        item_cost = sheet[f'I{row}'].value # Type = float
        line_item_value = quantity_ordered * item_cost
        line_item_value = round(line_item_value, 2)
        if order_number in po_value_dict:
            po_value_dict[order_number] = po_value_dict[order_number] + line_item_value
        else:
            po_value_dict[order_number] = line_item_value
    for key, value in po_value_dict.items():
        if key in current_currency_orders and value < int(min_order_value):
            orders_to_cancel_array.append(key)
    for row in range(4, sheet.max_row + 1):
        order_number = sheet[f'A{row}'].value
        if not order_number:
            break
        if order_number in orders_to_cancel_array:
            sheet[f'L{row}'].value = 0.0
            sheet[f'S{row}'].value = "CA - Cancelled: Not yet available"        
    for key in po_value_dict:
        if key in orders_to_cancel_array:
            po_value_dict[key] = 0.0
    workbook.save(new_file_path)
    return workbook, orders_to_cancel_array


def update_hand_off_date(workbook, new_file_path):
    sheet = workbook['Line Items']
    for row in range(4, sheet.max_row + 1):
        if not sheet[f'A{row}'].value:
            break
        else:
            sheet[f'R{row}'].value = sheet[f'P{row}'].value
    workbook.save(new_file_path)
    return workbook


def calculate_inventory(workbook, requested_inventory, orders_to_cancel_array):
    sheet = workbook['Line Items']
    for row in range(4, sheet.max_row + 1):
        order_number = sheet[f'A{row}'].value
        if not order_number:
            break
        model_number = sheet[f'C{row}'].value
        quantity_confirmed = int(float(sheet[f'L{row}'].value))
        if order_number not in orders_to_cancel_array: 
            if model_number in requested_inventory:
                requested_inventory[model_number] += quantity_confirmed
            else:
                requested_inventory[model_number] = quantity_confirmed
    requested_inventory = dict(sorted(requested_inventory.items()))
    return requested_inventory


def find_vendor_origins(submitted_files):
    if len(submitted_files) == 2:
        return "Amazon US + CA"
    elif "us" in submitted_files:
        return "Amazon US"
    else: 
        return "Amazon CA"


def copy_to_clipboard(inventory_dict, vendor_origins, is_confirmed):
    if not inventory_dict:
        messagebox.showinfo("Clipboard", "No requests - all orders below threshold or none requested.")
    else:
        if is_confirmed == False:
            clipboard_text = f"{vendor_origins} - Requested Items:\n"
        else: 
            clipboard_text = f"{vendor_origins} - Confirmed Items:\n"
        # Sort the inventory_dict by key (model) in alphabetical order
        sorted_inventory = sorted(inventory_dict.items())
        lines = [f"{model}: {quantity}" for model, quantity in sorted_inventory]
        clipboard_text += "\n".join(lines)
        pyperclip.copy(clipboard_text)
        messagebox.showinfo("Clipboard", "Inventory data copied to clipboard!")