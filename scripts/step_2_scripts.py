# step_2_scripts.py

import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.datetime import from_excel
import xlwt
import sys

def calculate_units_to_cancel(requested_inventory, available_inventory):
    units_to_cancel = {}
    for model_number in requested_inventory:
        if requested_inventory[model_number] > available_inventory[model_number]:
            discrepancy = requested_inventory[model_number] - available_inventory[model_number]
            units_to_cancel[model_number] = discrepancy    
    return units_to_cancel


def cancel_out_of_stock_units(units_to_cancel, available_inventory, processing_filenames, processing_dir, min_order_value_us, min_order_value_ca):
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = 'Combined Data'
    order_value_dict = {}
    current_row = 1
    
    # Create new temporary sheet with CA + USD data
    for filename in processing_filenames.values():
        processed_file_path = os.path.join(processing_dir, filename)
        processing_workbook = load_workbook(processed_file_path)
        reference_sheet = processing_workbook.active
        current_row = copy_data_to_new_sheet(reference_sheet, new_sheet, current_row)
    
    # Get order values
    order_value_dict = create_order_value_dict(new_sheet, order_value_dict)
    # First Pass: remove OOS units without putting orders below threshold (+ update order_value_dict)
    units_to_cancel, order_value_dict = remove_oos_units_above_threshold(new_sheet, units_to_cancel, order_value_dict, min_order_value_us, min_order_value_ca)
    # If more units need to be cancelled
    
    if units_to_cancel: 
        # sort order value dict low to high
        sorted_order_value_list = sorted(order_value_dict.items(), key=lambda item: item[1])
        # sort new sheet low to high
        sort_sheet_by_order_value(new_sheet, sorted_order_value_list, new_workbook)
        # Second Pass: iterate through temp sheet and remove remianing OOS units
        remove_oos_units_below_threshold(new_sheet, units_to_cancel)
        # Get updated order values
        order_value_dict = create_order_value_dict(new_sheet, order_value_dict)
        # Cancel orders below the minimum threshold
        cancel_orders_below_min_again(new_sheet, min_order_value_us, min_order_value_ca, order_value_dict)

    # Update the global inventory_dict (for the copy function)
    confirmed_inventory_dict = get_confirmed_inventory(new_sheet)

    # Check if there is any unallocated in-stock units (from orders which were cancelled in the second pass)
    extra_units_dict = get_extra_units(available_inventory, confirmed_inventory_dict)

    if extra_units_dict:
        # Get updated order values
        order_value_dict = create_order_value_dict(new_sheet, order_value_dict)
        # Add unallocated units to accepted orders
        allocate_extra_units_to_orders(new_sheet, extra_units_dict, order_value_dict)
        # Get the updated accepted_inventory_dict
        confirmed_inventory_dict = get_confirmed_inventory(new_sheet)

    # Create a new dict with order data from the new_sheet (order number, model number, units confirmed)
    final_order_data_dict = create_final_order_data_dict(new_sheet)

    # Update original sheet(s) with updated quantities + availability status
    update_original_sheets(processing_filenames, processing_dir, final_order_data_dict)

    # Save the new_sheet for debugging       
    new_file_path = os.path.join(processing_dir, 'Combined-Orders-Processing.xlsx')
    new_workbook.save(new_file_path)
    return confirmed_inventory_dict


def copy_data_to_new_sheet(reference_sheet, new_sheet, start_row):
    for row in range(4, reference_sheet.max_row + 1):
        if not reference_sheet[f'A{row}'].value:
            break
        for col in range(1, reference_sheet.max_column + 1):
            new_sheet.cell(row=start_row, column=col, value=reference_sheet.cell(row=row, column=col).value)
        start_row += 1
    return start_row


def create_order_value_dict(sheet, order_value_dict):
    order_value_dict = {}
    for row in range(1, sheet.max_row + 1):
        order_number = sheet[f'A{row}'].value
        list_price = sheet[f'I{row}'].value
        confirmed_qty = sheet[f'L{row}'].value
        cost = list_price * confirmed_qty
        if order_number in order_value_dict.keys():
            order_value_dict[order_number] = order_value_dict[order_number] + cost
        else:
            order_value_dict[order_number] = cost
    return order_value_dict


def remove_oos_units_above_threshold(sheet, units_to_cancel, order_value_dict, min_order_value_us, min_order_value_ca):
    min_order_value_us = float(min_order_value_us.get())
    min_order_value_ca = float(min_order_value_ca.get())
    for row in range(1, sheet.max_row + 1):
        if not units_to_cancel:
            break
        model_number = sheet[f'C{row}'].value
        if model_number in units_to_cancel.keys():
            cancel_amount = units_to_cancel[model_number]
            order_number = sheet[f'A{row}'].value
            price = sheet[f'I{row}'].value
            confirmed_qty = sheet[f'L{row}'].value
            currency = sheet[f'AF{row}'].value
            if currency == "USD":
                min_val = min_order_value_us
            else:
                min_val = min_order_value_ca
            while confirmed_qty > 0 and cancel_amount > 0:
                if (order_value_dict[order_number] - price) > min_val:
                    # confirmed_quantity: remove 1
                    sheet[f'L{row}'].value = sheet[f'L{row}'].value - 1
                    # update units_to_cancel
                    units_to_cancel[model_number] = units_to_cancel[model_number] - 1
                    # update order_value_dict
                    order_value_dict[order_number] = order_value_dict[order_number] - price
                    # update counters
                    confirmed_qty = confirmed_qty - 1
                    cancel_amount = cancel_amount - 1
                else:
                    break
            if units_to_cancel[model_number] == 0:
                del units_to_cancel[model_number]
    return units_to_cancel, order_value_dict
            

def sort_sheet_by_order_value(sheet, sorted_order_value_list, new_workbook):
    temp_sheet = new_workbook.create_sheet("TempSorted")
    new_row_idx = 1
    for order_number, _ in sorted_order_value_list:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
            if row[0].value == order_number:  # Assuming order_number is in the first column (column A)
                for col_idx, cell in enumerate(row, start=1):
                    temp_sheet.cell(row=new_row_idx, column=col_idx, value=cell.value)
                new_row_idx += 1
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    for row_idx, row in enumerate(temp_sheet.iter_rows(min_row=1, max_row=temp_sheet.max_row, values_only=False), start=1):
        for col_idx, cell in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell.value)
    new_workbook.remove(temp_sheet)


def remove_oos_units_below_threshold(sheet, units_to_cancel):
    for row in range(1, sheet.max_row + 1):
        order_number = sheet[f'A{row}'].value
        if not units_to_cancel:
            break
        model_number = sheet[f'C{row}'].value
        if model_number in units_to_cancel.keys():
            cancel_amount = units_to_cancel[model_number]
            confirmed_qty = sheet[f'L{row}'].value
            while confirmed_qty > 0 and cancel_amount > 0:
                # confirmed_quantity: remove 1
                sheet[f'L{row}'].value = sheet[f'L{row}'].value - 1
                # update units_to_cancel
                units_to_cancel[model_number] = units_to_cancel[model_number] - 1
                # update counters
                confirmed_qty = confirmed_qty - 1
                cancel_amount = cancel_amount - 1
            if units_to_cancel[model_number] == 0:
                del units_to_cancel[model_number]


def cancel_orders_below_min_again(sheet, min_order_value_us, min_order_value_ca, order_value_dict):
    for row in range(1, sheet.max_row + 1):
        if not sheet[f'A{row}'].value:
            break
        order_number = sheet[f'A{row}'].value
        model_number = sheet[f'C{row}'].value
        currency = sheet[f'AF{row}'].value
        if currency == "USD":
            minimum = int(min_order_value_us.get())
        else: 
            minimum = int(min_order_value_ca.get())
        if order_value_dict[order_number] < minimum:
            sheet[f'L{row}'].value = 0.0
            sheet[f'S{row}'].value = "CA - Cancelled: Not yet available"        


def get_confirmed_inventory(sheet):
    inventory_dict = {}
    for row in range(1, sheet.max_row + 1):
        if not sheet[f'A{row}'].value:
            break    
        model_number = sheet[f'C{row}'].value
        confirmed_qty = sheet[f'L{row}'].value
        if confirmed_qty != 0:
            if model_number in inventory_dict.keys():
                inventory_dict[model_number] = inventory_dict[model_number] + confirmed_qty
            else: 
                inventory_dict[model_number] = confirmed_qty
    return inventory_dict


def get_extra_units(available_inventory, confirmed_inventory_dict):
    extra_units_dict = {}
    for model in available_inventory.keys():
        if model in confirmed_inventory_dict.keys():
            variance = available_inventory[model] - confirmed_inventory_dict[model]
            if variance > 0:
                extra_units_dict[model] = variance
        else:
            extra_units_dict[model] = available_inventory[model]
    return extra_units_dict


def allocate_extra_units_to_orders(new_sheet, extra_units_dict, order_value_dict):
    for row in range(1, new_sheet.max_row + 1):
        order_number = new_sheet[f'A{row}'].value
        model_number = new_sheet[f'C{row}'].value
        requested_qty = new_sheet[f'J{row}'].value
        confirmed_qty = new_sheet[f'L{row}'].value
        if confirmed_qty < requested_qty:
            if model_number in extra_units_dict and order_value_dict.get(order_number, 0) > 0:
                while confirmed_qty < requested_qty and extra_units_dict[model_number] > 0:
                    new_sheet[f'L{row}'].value += 1
                    extra_units_dict[model_number] -= 1
                if extra_units_dict[model_number] == 0:
                    del extra_units_dict[model_number]
          

def create_final_order_data_dict(sheet):
    order_dict = {}
    for row in range(1, sheet.max_row + 1):
        order_number = sheet[f'A{row}'].value
        model_number = sheet[f'C{row}'].value
        confirmed_qty = sheet[f'L{row}'].value
        if order_number not in order_dict:
            order_dict[order_number] = {}
        order_dict[order_number][model_number] = confirmed_qty
    return order_dict


def update_original_sheets(processing_filenames, processing_dir, order_dict):
    for filename in processing_filenames.values():
        processed_file_path = os.path.join(processing_dir, filename)
        processing_workbook = load_workbook(processed_file_path)
        processing_sheet = processing_workbook.active
        for row in range(4, processing_sheet.max_row + 1):
            order_number = processing_sheet[f'A{row}'].value
            model_number = processing_sheet[f'C{row}'].value
            if order_number in order_dict and model_number in order_dict[order_number]:
                new_confirmed_qty = order_dict[order_number][model_number]
                processing_sheet[f'L{row}'].value = new_confirmed_qty
                if new_confirmed_qty == 0:
                    processing_sheet[f'S{row}'].value = "CA - Cancelled: Not yet available"
        processing_workbook.save(processed_file_path)


def convert_xlsx_to_xls(filename, processing_dir, upload_directory):
    processed_file_path = os.path.join(processing_dir, filename)
    processing_workbook = load_workbook(processed_file_path)
    reference_sheet = processing_workbook.active
    processing_workbook.close()
    completed_workbook = xlwt.Workbook()
    new_sheet = completed_workbook.add_sheet('Line Items')

    for row_idx, row in enumerate(reference_sheet.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.is_date:
                date_value = cell.value
                date_string = date_value.strftime("%d-%b-%Y")
                new_sheet.write(row_idx, col_idx, date_string)
            else:
                new_sheet.write(row_idx, col_idx, cell.value)
    
    # Require sheet - no need for contents, but it must be added or else Amazon will reject the file
    completed_workbook.add_sheet('Instructions')
    
    completed_filename = filename.replace('Processing.xlsx', 'Complete.xls')
    completed_file_path = os.path.join(upload_directory, completed_filename)
    completed_workbook.save(completed_file_path)
    
    print("processing_workbook: ", processing_workbook)
    print("processed_file_path: ", processed_file_path)
    return completed_workbook


